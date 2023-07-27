import openpyxl
import pandas as pd
from finance_model import ChartOfAccounts
import calendar
from datetime import datetime

MONTHS = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']


def read_trial_balance(year: int, month_number: int):
    print(f'year: month = {year} - {MONTHS[month_number]}')
    filename = f"HazTrain TB.{year} by month GENAESIS Confidential.xlsx"
    path = f'documents\\trial_balances\\{filename}'
    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    number_of_months = len(sheets)
    # if number_of_months < 12:
    #     print(f'Too few months: {year}')
    month = MONTHS[month_number]
    sheet = sheets[month_number]
    # print(f'month_num = {month_num} {year} {month}')
    if month not in sheet.lower() or str(year) not in sheet.lower():
        print(f'Not Found: {year} {month}: {sheet}')
        return None
    # tab = f'{month} {year} Trial Balance'
    df = pd.read_excel(path, sheet_name=month_number, na_values=0)
    df = df.replace(float('nan'), 0)
    return df


def clean_trial_balance(df: pd.DataFrame):
    if len(df.columns) != 4:
        raise Exception(f'error in clean_trial_balance: wrong number of columns: {len(df.columns)}')
    df = df.set_axis(["id", "description", "debit", "credit"], axis=1)
    df = df.astype({'id': int})
    df = df.set_index('id')
    df = df.drop([0])
    return df


def collapse_trail_balance(tb: pd.DataFrame, month, year):
    day = calendar.monthrange(year, month)[1]
    date = datetime(year, month, day)
    date_str = date.strftime('%Y-%m-%d')
    tb[date_str] = tb['debit'] - tb['credit']
    tb = tb[[date_str]]
    total = tb[date_str].sum()
    total = round(total, 6)
    if total != 0:
        raise Exception(f'ERROR: trail balances do not sum to zero: {total}')
    return tb


def read_all_trial_balances() -> pd.DataFrame:
    print('Finance model:')
    trial_balances = {}
    accounts = ChartOfAccounts()
    years = [2017, 2018, 2019, 2020, 2021, 2022, 2023]
    for year in years:
        print(f'year = {year}')
        filename = f"HazTrain TB.{year} by month GENAESIS Confidential.xlsx"
        path = f'documents\\trial_balances\\{filename}'
        wb = openpyxl.load_workbook(path)
        sheets = wb.sheetnames
        number_of_months = len(sheets)
        if number_of_months < 12:
            print(f'Too few months in {year}')
        for month_num in range(number_of_months):
            tb = read_trial_balance(year, month_num)
            tb = clean_trial_balance(tb)
            accounts.add_accounts(tb)
            tb = collapse_trail_balance(tb, month_num + 1, year)
            recs = tb.transpose().to_dict(orient='index')
            trial_balances = trial_balances | recs
        # return
        # break
    # print(trial_balances)
    df = pd.DataFrame(trial_balances).T
    df = df.replace(float('nan'), 0)
    cols = df.columns.tolist()
    cols.sort()
    return df, accounts


def main():
    tb, accounts = read_all_trial_balances()
    # accounts.write_accounts()


if __name__ == '__main__':
    # main()
    pass

# sheet = 'MARCH.2017 Trial Balance'
# code  = 'MARCH.2017 Trial Balance'
import matplotlib.pyplot as plt
import numpy as np

x = np.linspace(0, 2 * np.pi, 200)
y = np.sin(x)

fig, ax = plt.subplots()
ax.plot(x, y)
plt.show()